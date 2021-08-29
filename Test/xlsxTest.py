import openpyxl
import requests
from resource_data.keys import ServiceKey
import urllib
import math


# 한글을 URL로 넘겨주기 전에 인코딩 진행
def urlEncoding(val):
    encodeResult = urllib.parse.quote(val)
    return encodeResult


# 페이지와 검색어를 통해서 검색을 해주는 부분
def search(pageno , searKeword ):
    # 한글 인코딩 변경
    encodSearKeword = urlEncoding(searKeword)
    pageno = pageno
    dataJson = "http://api.visitkorea.or.kr/openapi/service/rest/PhotoGalleryService/gallerySearchList?keyword="+encodSearKeword+"&ServiceKey=" + ServiceKey + "&numOfRows=10&pageNo=" + str(pageno) + "&MobileOS=ETC&MobileApp=TestApp&_type=json"
    print(dataJson)
    response = requests.get(dataJson)
    data = response.json()
    print(data)
    
    # 검색 관광지 개수
    totalCount = data['response']['body']['totalCount']
    totalpage =  math.ceil(totalCount / 10)
    for i in range(1,totalpage+1):
        dataJson = "http://api.visitkorea.or.kr/openapi/service/rest/PhotoGalleryService/gallerySearchList?keyword=" + encodSearKeword + "&ServiceKey=" + ServiceKey + "&numOfRows=10&pageNo=" + str(i) + "&MobileOS=ETC&MobileApp=TestApp&_type=json"
        response = requests.get(dataJson)
        data = response.json()

    galTitle = data['response']['body']['items']['item'][0]['galTitle']
    galPhotographyLocation = data['response']['body']['items']['item'][0]['galPhotographyLocation']
    galSearchKeyword = data['response']['body']['items']['item'][0]['galSearchKeyword']
    galViewCount = data['response']['body']['items']['item'][0]['galViewCount']
    galWebImageUrl = data['response']['body']['items']['item'][0]['galWebImageUrl']
    makeExl(galTitle , galPhotographyLocation , galSearchKeyword , galViewCount , galWebImageUrl,searKeword)



def makeExl(galTitle, galPhotographyLocation, galSearchKeyword, galViewCount, galWebImageUrl,searKeword) :
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = searKeword
    sheet['B1'] = "galTitle"
    sheet['C1'] = "galPhotographyLocation"
    sheet['D1'] = "galSearchKeyword"
    sheet['E1'] = "galViewCount"
    sheet['F1'] = "galWebImageUrl"
    sheet.append(["",galTitle ,galPhotographyLocation, galSearchKeyword,galViewCount,galWebImageUrl])
    # 파일 저장
    wb.save("관광공사API정리.xlsx")


# xlsl 파일 생성
def xlslmake():
    wb = openpyxl.Workbook()
    wb.save("test.xlsx")
    
# xlsl 시트 수정
def xlsltest2():
    wb = openpyxl.Workbook()
    # 활성화 시트 불려오기
    sheet = wb.active
    # 시트 만들기
    sheet2 = wb.create_sheet('두번째 시트')
    # 시트 불러오기
    sheet2 = wb['두번째 시트']
    # 시트 이름 바꾸기
    sheet2.title = '수집데이터'

    wb.save("test.xlsx")

def test03():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append([1, "시작", 30, "안녕", 50])
    sheet['B2'] = 'bb2'
    sheet.cell(row=3, column=3).value = '3,3'
    sheet.append([1, 2, 3, 4, 5])
    sheet.append([10, 20, 30, 40, 50])

    wb.save('test2.xlsx')


def test03():
    wb = openpyxl.load_workbook('test.xlsx')
    sheet1 = wb.active
    sheet1.title = "이름111"
    sheet1.append(range(10))
    wb.save('test.xlsx')




if __name__ == '__main__':
    search(1, "경기도")



